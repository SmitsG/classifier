# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import openpyxl

def main():
    
    # Workbook1
    wb1 = Workbook()
    loaded_wb = wb1.load_workbook()
    sheet_names = wb1.get_sheet_names(loaded_wb)
    for sheet_name in sheet_names:
        sheet_name = Sheet(sheet_name)
        sheet_name.get_polarity()
        
class Workbook:
    def load_workbook(self):
        loaded_wb = openpyxl.load_workbook(filename = "C:/Users/Beheerder/Documents/GitHub/classifier/input.xlsx")
        return(loaded_wb)
        
    def get_sheet_names(self, wb):
        sheet_names = wb.get_sheet_names()
        return(sheet_names)
        
class Sheet:
    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        
    def get_polarity(self):
        pass
       
    def __str__(self):
        return(self.sheet_name)
        
        
class Aminoacid:
    def __init__(self, polarity, hydrofobicity, given, predicted, error):
        self.pol = polarity
        self.hyd = hydrofobicity
        self.giv = given
        self.pred = predicted
        self.err = error
        
    def get_polarity(self):
        return(self.pol)
    
    def get_hydrofobicity(self):
        return(self.hyd)
    
    def get_given(self):
        return(self.giv)
    
    def predicted(self):
        return(self.pred)
    
    def error(self):
        return(self.error)

    
main()